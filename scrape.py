from html import escape
from pathlib import Path
from urllib.parse import urlparse

import httpx
from bs4 import BeautifulSoup
from loguru import logger
from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parent
SEED_FILE = PROJECT_ROOT / "seeds" / "campus_restaurant_websites.xlsx"
HTML_DIR = PROJECT_ROOT / "html"
LOG_DIR = PROJECT_ROOT / "logs"
REQUEST_TIMEOUT_SECONDS = 5.0
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/135.0.0.0 Safari/537.36"
)


def setup_logging() -> Path:
    # Create the logs folder before adding the log file sink.
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    log_path = LOG_DIR / "scrape.log"

    logger.remove()
    logger.add(log_path, rotation="1 MB", encoding="utf-8")
    logger.add(lambda message: print(message, end=""))

    return log_path


def get_url_column_index(sheet) -> int:
    # Read the first row once and match the URL column without caring about case.
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))

    for index, value in enumerate(header_row):
        if isinstance(value, str) and value.strip().lower() == "url":
            return index

    raise ValueError("The Excel file does not contain a 'url' column.")


def read_urls_from_excel(file_path: Path) -> list[dict[str, str]]:
    # Load the workbook in read-only mode because we only need to read rows.
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    sheet = workbook.active
    url_index = get_url_column_index(sheet)
    rows: list[dict[str, str]] = []

    for row_number, row in enumerate(
        sheet.iter_rows(min_row=2, values_only=True),
        start=2,
    ):
        raw_url = row[url_index] if url_index < len(row) else None
        if not isinstance(raw_url, str) or not raw_url.strip():
            logger.warning(f"Skipping row {row_number} because the URL is empty.")
            continue

        rows.append({"row_number": str(row_number), "url": raw_url.strip()})

    workbook.close()
    return rows


def make_output_path(url: str, row_number: str) -> Path:
    # Build a stable file name from the row number and host name.
    parsed_url = urlparse(url)
    host_name = parsed_url.netloc.replace(":", "_") or "unknown-host"
    safe_host_name = "".join(
        character if character.isalnum() or character in {"-", "_", "."} else "-"
        for character in host_name
    )
    return HTML_DIR / f"row-{row_number}-{safe_host_name}.html"


def extract_readable_text(page_html: str) -> str:
    # Remove noisy tags and join visible text into readable lines.
    soup = BeautifulSoup(page_html, "html.parser")

    for tag in soup(["script", "style", "noscript"]):
        tag.decompose()

    lines = [text.strip() for text in soup.stripped_strings if text.strip()]
    return "\n".join(lines)


def save_readable_html(url: str, row_number: str, readable_text: str) -> Path:
    # Save the extracted text in a simple HTML wrapper for later review.
    HTML_DIR.mkdir(parents=True, exist_ok=True)
    output_path = make_output_path(url, row_number)

    html_content = "\n".join(
        [
            "<!DOCTYPE html>",
            "<html lang=\"en\">",
            "<head>",
            "  <meta charset=\"utf-8\">",
            f"  <title>Scraped page from {escape(url)}</title>",
            "</head>",
            "<body>",
            f"  <h1>{escape(url)}</h1>",
            "  <pre>",
            escape(readable_text),
            "  </pre>",
            "</body>",
            "</html>",
        ]
    )
    output_path.write_text(html_content, encoding="utf-8")

    return output_path


def fetch_and_save_pages(rows: list[dict[str, str]]) -> None:
    # Reuse one HTTP client so all requests share the same headers and timeout.
    headers = {"User-Agent": USER_AGENT}

    with httpx.Client(headers=headers, follow_redirects=True, timeout=REQUEST_TIMEOUT_SECONDS) as client:
        total_rows = len(rows)

        for index, row in enumerate(rows, start=1):
            url = row["url"]
            row_number = row["row_number"]
            logger.info(f"[{index}/{total_rows}] Fetching row {row_number}: {url}")

            try:
                response = client.get(url)
                response.raise_for_status()
                readable_text = extract_readable_text(response.text)

                if not readable_text:
                    logger.warning(f"No readable text found for row {row_number}: {url}")
                    continue

                output_path = save_readable_html(url, row_number, readable_text)
                logger.info(f"Saved row {row_number} to {output_path}")
            except httpx.TimeoutException:
                logger.error(f"Request timed out after 5 seconds for row {row_number}: {url}")
            except httpx.HTTPError as error:
                logger.error(f"HTTP error for row {row_number}: {url} | {error}")
            except Exception as error:
                logger.exception(f"Unexpected error for row {row_number}: {url} | {error}")


def main() -> None:
    # Run the full scrape pipeline from Excel input to saved HTML output.
    log_path = setup_logging()
    logger.info(f"Logging scraper activity to {log_path}")
    logger.info(f"Reading seed file: {SEED_FILE}")

    if not SEED_FILE.exists():
        raise FileNotFoundError(f"Seed file not found: {SEED_FILE}")

    rows = read_urls_from_excel(SEED_FILE)
    logger.info(f"Found {len(rows)} URLs to scrape.")
    fetch_and_save_pages(rows)
    logger.info("Scraping finished.")


if __name__ == "__main__":
    main()