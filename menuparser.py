import json
import os
from concurrent.futures import ThreadPoolExecutor, TimeoutError as FutureTimeoutError
from datetime import datetime
from html import unescape
from pathlib import Path
import time

from bs4 import BeautifulSoup
from dotenv import load_dotenv
from loguru import logger
from openai import OpenAI
from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parent
SEED_FILE = PROJECT_ROOT / "seeds" / "campus_restaurant_websites.xlsx"
HTML_DIR = PROJECT_ROOT / "html"
LOG_DIR = PROJECT_ROOT / "logs"
RESULTS_DIR = PROJECT_ROOT / "results"
MAX_HTML_TEXT_CHARS = 20000
CHUNK_CHAR_LIMIT = 6000
MODEL_TIMEOUT_SECONDS = 45
MODEL_MAX_RETRIES = 2


def setup_logging() -> Path:
    # Create the logs directory and configure both file and console logging.
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    log_path = LOG_DIR / "menuparser.log"

    logger.remove()
    logger.add(log_path, rotation="1 MB", encoding="utf-8")
    logger.add(lambda message: print(message, end=""))

    return log_path


def load_azure_settings() -> tuple[str, str, str]:
    # Load Azure OpenAI settings from .env so no key is hardcoded.
    load_dotenv()
    endpoint = os.getenv("AZURE_FOUNDRY_ENDPOINT")
    deployment = os.getenv("AZURE_FOUNDRY_MODEL")
    api_key = os.getenv("AZURE_FOUNDRY_API_KEY")

    missing = []
    if not endpoint:
        missing.append("AZURE_FOUNDRY_ENDPOINT")
    if not deployment:
        missing.append("AZURE_FOUNDRY_MODEL")
    if not api_key:
        missing.append("AZURE_FOUNDRY_API_KEY")

    if missing:
        raise ValueError(f"Missing required environment variables: {', '.join(missing)}")

    return endpoint, deployment, api_key


def get_column_indices(sheet) -> dict[str, int]:
    # Locate required columns once from the header row using case-insensitive matching.
    header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    required = {"url": None, "university": None, "restaurant": None}

    for index, value in enumerate(header):
        if not isinstance(value, str):
            continue

        key = value.strip().lower()
        if key == "url":
            required["url"] = index
        elif key == "university":
            required["university"] = index
        elif key in {"restaurant", "restaurant_name"}:
            required["restaurant"] = index

    missing = [name for name, index in required.items() if index is None]
    if missing:
        raise ValueError(f"Missing required Excel columns: {', '.join(missing)}")

    return {k: int(v) for k, v in required.items()}


def normalize_url(url: str) -> str:
    # Normalize URL text for stable lookup between seed and scraped files.
    return url.strip().rstrip("/").lower()


def load_metadata_by_url() -> dict[str, dict[str, str]]:
    # Build a URL-to-metadata map from the seed file for university and restaurant fields.
    workbook = load_workbook(SEED_FILE, read_only=True, data_only=True)
    sheet = workbook.active
    indices = get_column_indices(sheet)
    metadata_by_url: dict[str, dict[str, str]] = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        raw_url = row[indices["url"]] if indices["url"] < len(row) else None
        if not isinstance(raw_url, str) or not raw_url.strip():
            continue

        university = row[indices["university"]] if indices["university"] < len(row) else ""
        restaurant = row[indices["restaurant"]] if indices["restaurant"] < len(row) else ""

        metadata_by_url[normalize_url(raw_url)] = {
            "url": raw_url.strip(),
            "university": str(university or "").strip(),
            "restaurant_name": str(restaurant or "").strip(),
        }

    workbook.close()
    return metadata_by_url


def load_scraped_html(file_path: Path) -> tuple[str, str]:
    # Read scraped HTML and return source URL from <h1> plus visible text from <pre>.
    soup = BeautifulSoup(file_path.read_text(encoding="utf-8"), "html.parser")
    heading = soup.find("h1")
    pre_tag = soup.find("pre")

    url = unescape(heading.get_text(strip=True)) if heading else ""
    text = pre_tag.get_text("\n", strip=True) if pre_tag else ""

    if len(text) > MAX_HTML_TEXT_CHARS:
        logger.info(f"Truncating long page text to {MAX_HTML_TEXT_CHARS} characters: {file_path.name}")
        text = text[:MAX_HTML_TEXT_CHARS]

    return url, text


def safe_parse_json(message_text: str) -> list[dict]:
    # Parse JSON safely even when the model wraps output in markdown code fences.
    cleaned = message_text.strip()
    if cleaned.startswith("```"):
        cleaned = cleaned.strip("`")
        if cleaned.startswith("json"):
            cleaned = cleaned[4:].strip()

    try:
        parsed = json.loads(cleaned)
    except json.JSONDecodeError:
        logger.warning("Model response was not valid JSON. Skipping this chunk.")
        return []

    if not isinstance(parsed, list):
        return []
    return [item for item in parsed if isinstance(item, dict)]


def split_text_into_chunks(page_text: str, chunk_char_limit: int) -> list[str]:
    # Split long text into line-based chunks so each model request stays manageable.
    if len(page_text) <= chunk_char_limit:
        return [page_text]

    chunks: list[str] = []
    current_lines: list[str] = []
    current_size = 0

    for line in page_text.splitlines():
        line_size = len(line) + 1
        if current_lines and current_size + line_size > chunk_char_limit:
            chunks.append("\n".join(current_lines))
            current_lines = []
            current_size = 0

        current_lines.append(line)
        current_size += line_size

    if current_lines:
        chunks.append("\n".join(current_lines))

    return chunks


def normalize_meal_type(meal_type: str) -> str:
    # Convert meal type to one of the allowed values.
    value = str(meal_type or "").strip().lower()
    if value in {"breakfast", "lunch", "dinner", "unknown"}:
        return value
    return "unknown"


def is_valid_menu_item(item: dict) -> bool:
    # Skip items that are empty, invalid, or low-confidence according to model output.
    meal_name = str(item.get("meal_name", "")).strip()
    is_valid = bool(item.get("is_valid_menu", False))
    confidence = float(item.get("name_confidence", 0.0) or 0.0)

    if not meal_name:
        return False
    if not is_valid:
        return False
    if confidence < 0.7:
        return False

    return True


def build_output_item(
    extracted_item: dict,
    source_url: str,
    university: str,
    restaurant_name: str,
) -> dict:
    # Build final JSON object using required keys and safe fallback values.
    scrape_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    menu_date = str(extracted_item.get("menu_date", "")).strip()
    meal_type = normalize_meal_type(str(extracted_item.get("meal_type", "unknown")))
    meal_name = str(extracted_item.get("meal_name", "")).strip()
    price_krw = str(extracted_item.get("price_krw", "")).strip()

    return {
        "scrape_date": scrape_date,
        "url": source_url,
        "menu_date": menu_date,
        "meal_type": meal_type,
        "meal_name": meal_name,
        "price_krw": price_krw,
        "university": university,
        "restaurant_name": restaurant_name,
    }


def extract_menu_with_ai(client: OpenAI, model_name: str, source_url: str, page_text: str) -> list[dict]:
    # Ask the model for structured menu extraction and return parsed list output.
    prompt = f"""
You are extracting cafeteria menu items.

Return ONLY a JSON array. Each element must include these keys:
- menu_date: string in YYYY-MM-DD when available, else ""
- meal_type: breakfast, lunch, dinner, or unknown
- meal_name: food name only
- price_krw: number-like string without commas (example: "5000") or ""
- is_valid_menu: true only when meal_name is clearly a real meal item
- name_confidence: number between 0 and 1 for meal_name confidence

Rules:
- Skip non-menu text, notices, navigation labels, and random words.
- If unsure whether meal_name is a real menu item, set is_valid_menu to false.
- If price is missing, set price_krw to "".
- Do not include any keys other than the required keys above.

Source URL:
{source_url}

Scraped Text:
{page_text}
"""

    response = client.chat.completions.create(
        model=model_name,
        messages=[
            {"role": "system", "content": "You are a careful data extraction assistant."},
            {"role": "user", "content": prompt},
        ],
    )

    content = response.choices[0].message.content or "[]"
    return safe_parse_json(content)


def extract_menu_with_retries(client: OpenAI, model_name: str, source_url: str, chunk_text: str) -> list[dict]:
    # Retry model calls for transient API issues and continue even if one attempt fails.
    for attempt in range(1, MODEL_MAX_RETRIES + 1):
        executor = ThreadPoolExecutor(max_workers=1)
        future = executor.submit(extract_menu_with_ai, client, model_name, source_url, chunk_text)

        try:
            result = future.result(timeout=MODEL_TIMEOUT_SECONDS)
            executor.shutdown(wait=False, cancel_futures=True)
            return result
        except FutureTimeoutError:
            logger.warning(
                f"Model call timed out after {MODEL_TIMEOUT_SECONDS}s "
                f"(attempt {attempt}/{MODEL_MAX_RETRIES}) for {source_url}"
            )
        except Exception as error:
            logger.warning(f"Model call failed (attempt {attempt}/{MODEL_MAX_RETRIES}) for {source_url}: {error}")
        finally:
            executor.shutdown(wait=False, cancel_futures=True)

        if attempt < MODEL_MAX_RETRIES:
            time.sleep(2)

    return []


def deduplicate_items(items: list[dict]) -> list[dict]:
    # Remove repeated menu records gathered across chunks from the same page.
    seen_keys: set[tuple[str, str, str, str]] = set()
    unique_items: list[dict] = []

    for item in items:
        key = (
            str(item.get("menu_date", "")).strip(),
            str(item.get("meal_type", "")).strip().lower(),
            str(item.get("meal_name", "")).strip(),
            str(item.get("price_krw", "")).strip(),
        )
        if key in seen_keys:
            continue

        seen_keys.add(key)
        unique_items.append(item)

    return unique_items


def append_jsonl(file_path: Path, item: dict) -> None:
    # Write one JSON record per line as soon as each valid item is available.
    with file_path.open("a", encoding="utf-8") as output_file:
        output_file.write(json.dumps(item, ensure_ascii=False) + "\n")


def run_parser() -> Path:
    # Execute the full parsing pipeline from scraped HTML files to JSONL output.
    endpoint, deployment, api_key = load_azure_settings()
    # Configure timeout directly on the client so each API call cannot hang forever.
    client = OpenAI(base_url=endpoint, api_key=api_key, timeout=MODEL_TIMEOUT_SECONDS)
    metadata_by_url = load_metadata_by_url()
    html_files = sorted(HTML_DIR.glob("*.html"))

    RESULTS_DIR.mkdir(parents=True, exist_ok=True)
    output_path = RESULTS_DIR / f"menus-{datetime.now().strftime('%Y-%m-%d')}.jsonl"

    logger.info(f"Found {len(html_files)} scraped HTML files.")
    logger.info(f"Saving JSONL records to {output_path}")

    for index, html_file in enumerate(html_files, start=1):
        logger.info(f"[{index}/{len(html_files)}] Parsing {html_file.name}")

        try:
            source_url, page_text = load_scraped_html(html_file)
            if not source_url or not page_text:
                logger.warning(f"Skipping {html_file.name} because URL or text is missing.")
                continue

            meta = metadata_by_url.get(normalize_url(source_url), {})
            university = meta.get("university", "")
            restaurant_name = meta.get("restaurant_name", "")

            text_chunks = split_text_into_chunks(page_text, CHUNK_CHAR_LIMIT)
            logger.info(f"Split page into {len(text_chunks)} chunk(s): {source_url}")

            extracted_items: list[dict] = []
            for chunk_index, chunk_text in enumerate(text_chunks, start=1):
                logger.info(f"Extracting chunk {chunk_index}/{len(text_chunks)} for {html_file.name}")
                chunk_items = extract_menu_with_retries(client, deployment, source_url, chunk_text)
                extracted_items.extend(chunk_items)

            extracted_items = deduplicate_items(extracted_items)
            logger.info(f"Model returned {len(extracted_items)} unique raw items for {source_url}")

            valid_count = 0
            for extracted_item in extracted_items:
                if not is_valid_menu_item(extracted_item):
                    continue

                output_item = build_output_item(
                    extracted_item=extracted_item,
                    source_url=source_url,
                    university=university,
                    restaurant_name=restaurant_name,
                )

                append_jsonl(output_path, output_item)
                valid_count += 1

            logger.info(f"Saved {valid_count} valid menu items from {html_file.name}")
        except Exception as error:
            logger.exception(f"Failed to parse {html_file.name}: {error}")

    return output_path


def main() -> None:
    # Configure logging and run menu parsing with clear start and end messages.
    log_path = setup_logging()
    logger.info(f"Logging menu parser activity to {log_path}")

    if not SEED_FILE.exists():
        raise FileNotFoundError(f"Seed file not found: {SEED_FILE}")
    if not HTML_DIR.exists():
        raise FileNotFoundError(f"Scraped HTML directory not found: {HTML_DIR}")

    output_path = run_parser()
    logger.info(f"Menu parsing finished. Output file: {output_path}")


if __name__ == "__main__":
    main()